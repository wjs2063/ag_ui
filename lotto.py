"""로또 번호 생성기 - 연속 숫자 1쌍 이상 포함, 엑셀 출력"""

import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 로또 공 색상 (번호 구간별)
BALL_COLORS = {
    range(1, 11): "FFEAA7",   # 노랑
    range(11, 21): "74B9FF",  # 파랑
    range(21, 31): "FF7675",  # 빨강
    range(31, 41): "A29BFE",  # 보라
    range(41, 46): "55EFC4",  # 초록
}


def get_ball_color(num: int) -> str:
    for r, color in BALL_COLORS.items():
        if num in r:
            return color
    return "FFFFFF"


def pick_lotto() -> list[int]:
    """1~45 중 6개, 연속된 숫자 쌍이 1개 이상 포함"""
    while True:
        nums = sorted(random.sample(range(1, 46), 6))
        if any(nums[i + 1] - nums[i] == 1 for i in range(5)):
            return nums


def find_consecutive_pairs(nums: list[int]) -> list[tuple[int, int]]:
    """연속된 숫자 쌍 찾기"""
    return [
        (nums[i], nums[i + 1])
        for i in range(len(nums) - 1)
        if nums[i + 1] - nums[i] == 1
    ]


def generate_excel(count: int, filename: str = "lotto.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "로또번호"

    # ── 스타일 ──
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3436")
    cell_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    consec_font = Font(bold=True, size=11, color="D63031")

    # ── 헤더 ──
    headers = ["회차", "번호1", "번호2", "번호3", "번호4", "번호5", "번호6", "연속쌍"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # ── 데이터 ──
    for i in range(count):
        row = i + 2
        nums = pick_lotto()
        pairs = find_consecutive_pairs(nums)

        ws.cell(row=row, column=1, value=i + 1).alignment = center
        ws.cell(row=row, column=1).border = thin_border

        for j, num in enumerate(nums):
            cell = ws.cell(row=row, column=j + 2, value=num)
            cell.font = cell_font
            cell.alignment = center
            cell.border = thin_border
            cell.fill = PatternFill("solid", fgColor=get_ball_color(num))

            # 연속 숫자는 빨간 글씨로 강조
            if any(num in pair for pair in pairs):
                cell.font = consec_font

        # 연속쌍 표시
        pair_str = ", ".join(f"{a}-{b}" for a, b in pairs)
        cell = ws.cell(row=row, column=8, value=pair_str)
        cell.alignment = center
        cell.border = thin_border
        cell.font = Font(color="D63031")

    # ── 열 너비 ──
    ws.column_dimensions["A"].width = 8
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions["H"].width = 18

    # ── 저장 ──
    output = Path(filename)
    wb.save(output)
    print(f"{count}회 생성 완료 → {output.resolve()}")


if __name__ == "__main__":
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 5
    generate_excel(n)
