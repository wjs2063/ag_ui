"""
당첨번호(6개 + 보너스 1개)를 입력하면, picks/ 폴더 안의 모든 엑셀파일
(Combined 시트)에 들어 있는 조합들이 각각 몇 등인지 집계한다.

등수
  1등: 6개 일치
  2등: 5개 일치 + 보너스 1개 일치
  3등: 5개 일치
  4등: 4개 일치
  5등: 3개 일치
  그 외: 낙첨

사용 예시
  $ python3 check_rank.py 5 10 12 27 31 33 --bonus 1
  $ python3 check_rank.py 5,10,12,27,31,33 -b 1
  $ python3 check_rank.py 5 10 12 27 31 33 -b 1 --dir picks
  $ python3 check_rank.py 5 10 12 27 31 33 -b 1 --show-all
"""

import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).parent
PICKS_DIR_DEFAULT = BASE_DIR / "picks"

RANK_LABEL = {
    1: "1등 (6개)",
    2: "2등 (5개+보너스)",
    3: "3등 (5개)",
    4: "4등 (4개)",
    5: "5등 (3개)",
}


def parse_numbers(values):
    nums = []
    for v in values:
        for x in str(v).replace(",", " ").split():
            nums.append(int(x))
    return nums


def rank_of(pick, winning_set, bonus):
    matched = len(set(pick) & winning_set)
    if matched == 6:
        return 1
    if matched == 5 and bonus in pick:
        return 2
    if matched == 5:
        return 3
    if matched == 4:
        return 4
    if matched == 3:
        return 5
    return None


def load_picks_from_dir(picks_dir: Path):
    """picks_dir 안의 모든 .xlsx 의 Combined 시트에서 (파일, 순번, 버전키, pick) 로드."""
    rows = []
    files = []
    for path in sorted(picks_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [경고] {path.name} 열기 실패: {e}")
            continue
        if "Combined" not in wb.sheetnames:
            wb.close()
            print(f"  [스킵] {path.name}: Combined 시트 없음")
            continue
        ws = wb["Combined"]
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            idx, key = row[0], row[1]
            nums = row[2:8]
            if any(x is None for x in nums):
                continue
            rows.append((path.name, idx, key, tuple(sorted(int(x) for x in nums))))
            n += 1
        wb.close()
        files.append((path.name, n))
    return rows, files


def main():
    parser = argparse.ArgumentParser(description="picks 폴더 조합 등수 판정")
    parser.add_argument("numbers", nargs="*",
                        help="당첨번호 6개 (공백 또는 콤마 구분)")
    parser.add_argument("-b", "--bonus", type=int, required=True, help="보너스 번호")
    parser.add_argument("--dir", type=str, default=None,
                        help=f"조합 엑셀이 들어 있는 폴더 (기본: {PICKS_DIR_DEFAULT.name}/)")
    parser.add_argument("--show-all", action="store_true",
                        help="등수에 들지 못한 조합도 모두 출력")
    args = parser.parse_args()

    winning = parse_numbers(args.numbers)
    if len(winning) != 6:
        parser.error(f"당첨번호 6개를 입력하세요 (입력: {len(winning)}개)")
    if len(set(winning)) != 6:
        parser.error("당첨번호 6개에 중복이 있습니다.")
    if args.bonus in winning:
        parser.error("보너스 번호가 당첨번호와 겹칩니다.")
    for n in (*winning, args.bonus):
        if not 1 <= n <= 45:
            parser.error(f"번호 {n} 은 1~45 범위가 아닙니다.")

    picks_dir = Path(args.dir).expanduser() if args.dir else PICKS_DIR_DEFAULT
    if not picks_dir.is_absolute():
        picks_dir = (BASE_DIR / picks_dir).resolve()
    if not picks_dir.is_dir():
        parser.error(f"폴더 없음: {picks_dir}")

    winning_sorted = sorted(winning)
    winning_set = set(winning)

    print(f"[당첨번호] {winning_sorted}  (보너스 {args.bonus})")
    print(f"[대상 폴더] {picks_dir}")

    rows, files = load_picks_from_dir(picks_dir)
    if not files:
        print("  (폴더에 Combined 시트를 가진 엑셀파일이 없습니다.)")
        return

    print(f"[로드] 파일 {len(files)}개  ·  총 조합 {len(rows)}건")
    for name, n in files:
        print(f"    - {name}  ({n}건)")

    # 전체 집계
    counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, None: 0}
    # 파일별 집계
    per_file = defaultdict(lambda: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, None: 0})
    winners = []
    for fname, idx, key, pick in rows:
        r = rank_of(pick, winning_set, args.bonus)
        counts[r] += 1
        per_file[fname][r] += 1
        if r is not None:
            winners.append((r, fname, idx, key, pick))

    print()
    print("=" * 70)
    print(" 전체 등수별 집계")
    print("=" * 70)
    for r in (1, 2, 3, 4, 5):
        print(f"  {RANK_LABEL[r]:<20} : {counts[r]:5d}건")
    print(f"  {'낙첨':<20} : {counts[None]:5d}건")
    total_winners = sum(counts[r] for r in (1, 2, 3, 4, 5))
    print(f"  {'-' * 30}")
    print(f"  {'당첨 합계':<20} : {total_winners:5d}건  /  {len(rows)}건")

    print()
    print("=" * 70)
    print(" 파일별 집계")
    print("=" * 70)
    header = f"  {'파일명':<42} {'1등':>4} {'2등':>4} {'3등':>4} {'4등':>4} {'5등':>4} {'낙첨':>5}"
    print(header)
    for name, _ in files:
        c = per_file[name]
        print(f"  {name[:42]:<42} {c[1]:>4} {c[2]:>4} {c[3]:>4} {c[4]:>4} {c[5]:>4} {c[None]:>5}")

    if winners:
        print()
        print("=" * 70)
        print(" 당첨 조합 상세")
        print("=" * 70)
        winners.sort(key=lambda x: (x[0], x[1], x[2]))
        for r, fname, idx, key, pick in winners:
            matched = sorted(set(pick) & winning_set)
            bonus_mark = " +B" if r == 2 else ""
            print(f"  {r}등  [{fname}] 순번 {idx:4d}  [{key:<22}]  "
                  f"{' '.join(f'{n:2d}' for n in pick)}  매치={matched}{bonus_mark}")

    if args.show_all:
        print()
        print("=" * 70)
        print(" 전체 조합 (매치 개수)")
        print("=" * 70)
        for fname, idx, key, pick in rows:
            matched = len(set(pick) & winning_set)
            r = rank_of(pick, winning_set, args.bonus)
            tag = f"  → {r}등" if r is not None else ""
            print(f"  [{fname}] {idx:4d}  [{key:<22}]  "
                  f"{' '.join(f'{n:2d}' for n in pick)}  매치={matched}{tag}")


if __name__ == "__main__":
    main()
