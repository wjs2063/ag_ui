"""
past_lotto.xlsx 의 역대 당첨 데이터를 기반으로 로또 번호를 추첨한다.

기본 제외 번호: {1, 2, 16, 18, 42, 43, 44, 45}
필수 필터:
  (a) 위 8개 번호는 후보에서 영구 제외 (--exclude-numbers 로 덮어쓸 수 있음)
  (b) 과거 1등 당첨 조합과 동일한 픽은 폐기
  (c) 3개 이상 연속된 숫자(예: 5-6-7) 가 포함된 픽은 폐기

============================================================
당첨 확률
============================================================
- 한국 로또 1등은 본질적으로 C(45, 6) = 8,145,060 가지 중 하나를 맞춰야 한다.
  → 무작위 단일 티켓 기준 수학적 1등 확률 = 1 / 8,145,060 ≈ 0.00001228 % (약 1/814만)

- 본 알고리즘이 만드는 후보 모집단(pool) 크기 (위 필수 필터 적용 후):
    · 8개 번호 제외 후 단순 6조합:           C(37, 6)        = 2,324,784
    · 위에서 3개 이상 연속 케이스 추가 제거: ≈ 2,149,548
    · 추가로 과거 1등 당첨 조합(허용 번호만 사용한 회차)도 제거 → ~2,149,3xx
- 즉 실제 추첨 결과(45개 중 6개 무작위)가 본 알고리즘 풀 안에 있을 확률은
    pool / 8,145,060 ≈ 2,149,548 / 8,145,060 ≈ 26.4 %
  바꿔 말하면, 실제 1등 조합의 약 73.6 % 는 본 알고리즘이 아예 만들지 못한다.

- 단일 티켓의 1등 당첨 확률 (수학적 결론):
    P(win) = (pool / 8,145,060) × (1 / pool) = 1 / 8,145,060 ≈ 0.00001228 %
  → 풀을 좁혀도 "단일 티켓"의 수학적 1등 확률은 변하지 않는다.
    (필터는 통계적으로 흔치 않은 패턴을 버려 *체감* 적중률을 높이는 휴리스틱일 뿐이다.)

- 참고: N 세트를 동시에 사면 P(적어도 1세트 1등) ≈ N / 8,145,060.
    예) 50세트 → ≈ 1 / 162,901 ≈ 0.000614 %

============================================================
사용법
============================================================

1) 의존성 설치
   $ pip3 install openpyxl

2) 기본 실행 (모든 분석 버전을 각 5세트씩 출력)
   $ python3 lotto_generator.py
   - 자동 적용되는 필수 필터:
       (a) 후보에서 {1, 2, 16, 18, 42, 43, 44, 45} 영구 제외
       (b) 과거 1등 당첨과 동일한 조합 폐기
       (c) 3개 이상 연속된 숫자(예: 5-6-7) 포함 픽 폐기
       (d) 모든 버전 합쳐 동일한 6조합이 두 번 나오지 않음 (버전 간 중복 제거)
   - 콘솔 첫 줄에 "[기본 제외] ... (남은 후보 37개)" 가 표시되면 정상 동작

3) 버전별 세트 수 지정
   $ python3 lotto_generator.py -c 10        # 버전당 10세트
   $ python3 lotto_generator.py --count 3

4) 특정 버전만 실행
   $ python3 lotto_generator.py --versions hot-one-consecutive uniform-no-consecutive
   사용 가능한 키: hot-no-consecutive, hot-one-consecutive, hot-two-consecutive,
                  cold-no-consecutive, cold-one-consecutive,
                  uniform-no-consecutive, uniform-one-consecutive,
                  zone-no-consecutive, zone-one-consecutive, zone-two-consecutive

5) 재현 가능한 결과 (디버그/공유용)
   $ python3 lotto_generator.py --seed 42

6) 빈도/확률 표 생략
   $ python3 lotto_generator.py --no-stats

7) 가중치/제외 동작 자체를 검증
   $ python3 lotto_generator.py --verify
   - 과거 당첨 조합이 결과에 절대 포함되지 않는지
   - 가중치 모드별 샘플링이 의도된 분포를 따르는지 확인

8) 결과를 엑셀로 저장 (Frequency / Summary / 버전별 시트)
   - 기본 동작: 매 실행마다 자동으로 저장된다 (--xlsx 안 줘도 저장).
   $ python3 lotto_generator.py -c 10
        → picks/lotto__excl-1-2-16-18-42-43-44-45__20260510-143025.xlsx
   $ python3 lotto_generator.py --versions hot-one-consecutive --xlsx hot1
        → picks/hot1__excl-1-2-16-18-42-43-44-45__20260510-143025.xlsx
   - 저장 위치는 스크립트 옆의 picks/ 폴더로 고정 (없으면 자동 생성).
     --xlsx 에 디렉터리를 적어도 무시되고 파일명 stem 만 사용된다.
   - 자동 파일명 규칙:
     "picks/{stem}__excl-{제외번호들}__{YYYYMMDD-HHMMSS}.xlsx"
   - 같은 명령을 다시 돌려도 시각이 달라 덮어쓰기 사고가 나지 않는다.
   - 매 실행 시작할 때 picks/ 안의 모든 기존 xlsx 를 스캔해 그 안의 6조합을
     중복 방지 풀에 미리 채운다 → 파일이 여러 개여도 picks/ 전체에서
     동일한 6조합이 두 번 나오지 않는다.
   - 저장을 끄고 싶다면 --no-save:
   $ python3 lotto_generator.py -c 5 --no-save

9) 종합 모드: 모든 버전을 합쳐 총 N개 (각 버전에 균등 분배)
   $ python3 lotto_generator.py --total 500 --xlsx combined.xlsx
   - --total 지정시 -c 는 무시되며, Combined 시트가 추가로 생성됨
   - --versions 와 함께 쓰면 선택한 버전들에만 균등 분배

10) 구간 분할 개수 변경 (기본 7구간, 7~8 권장)
   $ python3 lotto_generator.py --zones 8 --total 500 --xlsx zone8.xlsx

11) "구간확률 적용" 버전(zone-*) 만으로 500개 추첨
   $ python3 lotto_generator.py --zones 7 --total 500 \
       --versions zone-no-consecutive zone-one-consecutive zone-two-consecutive \
       --xlsx zone_only_500.xlsx --no-stats

12) 특정 번호들이 동시에 등장하면 폐기 (반복 지정 가능)
   $ python3 lotto_generator.py --total 500 --exclude-pair 1,2 --xlsx out.xlsx
   $ python3 lotto_generator.py --total 500 --exclude-pair 1,2 --exclude-pair 44,45

13) 후보 번호 자체를 영구 제외 (빈도/확률/구간 분할에서도 빠짐)
   $ python3 lotto_generator.py --exclude-numbers 1,2,45 --zones 6 --total 500 \
       --versions zone-no-consecutive zone-one-consecutive zone-two-consecutive \
       --no-stats --xlsx zone_only_500.xlsx
   - 1,2,45 제외 → 남은 42개를 6구간 = 한 구간 7개 폭으로 분할
   - zone 모드는 이 새 구간 패턴 분포에 따라 추첨
   - 주의: --exclude-numbers 를 직접 지정하면 위 2)의 기본 제외 8개는 적용되지 않음.
     기본값을 유지하면서 추가 제외하려면 모두 나열해야 한다.
       예) $ python3 lotto_generator.py --exclude-numbers 1,2,16,18,42,43,44,45,7

14) 기본 제외 번호를 끄고 1~45 전부 후보로 쓰기
   $ python3 lotto_generator.py --exclude-numbers 0
   - 존재하지 않는 번호 하나만 넘기면 EXCLUDED 가 사실상 비어 있게 되어
     1~45 전체가 후보가 됨 (3개 이상 연속/과거 당첨 필터는 그대로 유지)

15) 버전 간 중복 없이 N세트 뽑아 엑셀로 저장
   $ python3 lotto_generator.py --total 200 --xlsx unique200.xlsx
   - 모든 버전이 같은 'seen' 풀을 공유 → Combined 시트의 200행은 모두 서로 다른 6조합
   - 한 버전 안 / 버전 간 어디서도 동일한 6조합이 두 번 등장하지 않음

16) "N번 돌리고 그때 안 나온 케이스로 뽑기" — 워밍업 여집합 추첨
   $ python3 lotto_generator.py --warmup 10000 --total 50
   - 알고리즘 유니버스(약 2.1M) 중에서 먼저 10,000개를 미리 뽑아 seen 에만 추가 → 폐기
   - 그 다음 50개를 뽑되, 위 10,000개와 picks/ 안 기존 픽들과 절대 안 겹침
   - 즉 출력 50개는 (유니버스) − (워밍업 10,000 ∪ 기존 picks/) 의 여집합에서 선택됨
   - 워밍업 픽 자체는 저장되지 않으며 콘솔에도 출력되지 않는다
   - 같은 명령을 반복하면 picks/ 누적 + 워밍업 N 이 합쳐져 점점 새로운 영역으로 밀려난다

17) "X개 뽑아 저장" 한 묶음을 N번 반복 — 라운드 누적 여집합 추첨
   $ python3 lotto_generator.py --total 50 --repeat 200
   - 1라운드: 50개 뽑기 → picks/...__r001.xlsx 저장
   - 2라운드: 1라운드의 50개를 피해 새 50개 → ..._r002.xlsx 저장
     ...
   - 200라운드: 1~199라운드의 9,950개를 모두 피해 새 50개 → ..._r200.xlsx 저장
   - 즉 마지막 파일(_r200) 이 곧 "199번 돌리고 그때 나온 케이스를 모두 제외한 여집합"
     에서 뽑힌 결과이며, 200개 파일 합계 10,000세트는 모두 서로 다른 6조합.
   - --warmup 와 함께 쓰면 워밍업 → 라운드 1 → ... → 라운드 N 순으로 모두 누적 제외.
   - 라운드 모드에서는 콘솔 미리보기는 생략되고, 라운드별 1줄 요약만 출력된다.

18) 마지막 여집합 라운드만 더 많이 뽑기 (--final-count)
   $ python3 lotto_generator.py --total 50 --repeat 200 --final-count 1000
   - r001 ~ r199 : 각 50개씩 (앞 199라운드 누적 9,950개)
   - r200        : 1,000개 (앞 9,950개를 모두 피한 여집합에서 추출)
   - 마지막 파일(_r200) 이 1000세트짜리 "여집합 추첨 결과" 가 된다.
   - --final-count 단독(=--repeat 없음) 은 의미 없음 (마지막 라운드 = 첫 라운드).

============================================================
알고리즘
============================================================
- number_frequency()   : 1~45 각 번호가 과거 1등에 등장한 횟수
- number_probability() : 빈도 / 전체 = 출현 확률 (합 = 1.0)
- build_weights()      : 위 확률 또는 그 변형(역가중/균등)을 샘플링 가중치로
- weighted_sample_unique(): 가중치를 누적분포로 사용해 6개 비복원 추출

- ZONES                : 1-10 / 11-20 / 21-30 / 31-40 / 41-45 (5구간)
- zone_distribution()  : 과거 회차의 구간별 등장 합 + 구간 카운트 패턴 빈도
- adjacent_gaps()      : sorted 6개의 인접 차이 5개
- gap_distribution()   : 과거 모든 인접 간격 빈도

- zone-mode 추출:
    1) 과거 zone 패턴 (예: (2,1,2,1,0)) 분포에서 한 패턴을 확률에 비례해 추첨
    2) 각 zone 안에서는 번호별 출현확률 가중치로 그 zone 카운트만큼 비복원 추출
    3) 결과가 과거 당첨/연속쌍 조건 불일치면 재시도

- generate_pick()      : 모든 모드 공통 — (a) 과거 당첨 조합이면 폐기,
                         (b) 연속쌍 조건 불일치면 폐기,
                         (c) 3개 이상 연속된 숫자(1-2-3 등)가 있으면 모드와 무관하게 폐기,
                         (d) 모두 만족시 채택
"""

import argparse
import random
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "past_lotto.xlsx"
PICKS_DIR = Path(__file__).parent / "picks"
NUMBER_RANGE = range(1, 46)
PICK_SIZE = 6
MAX_ATTEMPTS = 200_000

# 알고리즘 기본 제외 번호. CLI --exclude-numbers 로 덮어쓰지 않으면 항상 적용됨.
DEFAULT_EXCLUDED_NUMBERS = frozenset({1, 2, 16, 18, 42, 43, 44, 45})

# 후보 번호 풀 (글로벌). --exclude-numbers 로 갱신됨.
EXCLUDED_NUMBERS = set(DEFAULT_EXCLUDED_NUMBERS)
ALLOWED_NUMBERS = sorted(set(NUMBER_RANGE) - EXCLUDED_NUMBERS)


def make_zones(n_zones, allowed=None):
    """allowed 정렬 리스트를 n_zones 구간으로 거의 균등 분할.
    각 zone은 (lo, hi, members) 형태.

    예) allowed=range(1,46), n_zones=7 → (1-7, 8-14, ..., 40-45)
        allowed=range(3,45) (42개), n_zones=6 → (3-9, 10-16, 17-23, 24-30, 31-37, 38-44)
    """
    if allowed is None:
        allowed = list(NUMBER_RANGE)
    allowed = sorted(allowed)
    n = len(allowed)
    if n_zones < 1 or n_zones > n:
        raise ValueError(f"n_zones must be in 1..{n}")
    base, rem = divmod(n, n_zones)
    zones = []
    idx = 0
    for i in range(n_zones):
        size = base + (1 if i < rem else 0)
        members = allowed[idx:idx + size]
        zones.append((members[0], members[-1], members))
        idx += size
    return zones


# 기본값: 1~45 / 7구간
ZONES = make_zones(7, ALLOWED_NUMBERS)
ZONE_BY_N = {n: zi for zi, (_, _, members) in enumerate(ZONES) for n in members}


def _refresh_zone_index():
    """ZONES 가 바뀐 뒤 ZONE_BY_N 재계산."""
    global ZONE_BY_N
    ZONE_BY_N = {n: zi for zi, (_, _, members) in enumerate(ZONES) for n in members}


def zone_index(n):
    if n not in ZONE_BY_N:
        raise ValueError(f"숫자 {n} 은 어떤 구간에도 속하지 않음 (제외된 번호일 수 있음)")
    return ZONE_BY_N[n]


def zone_pattern(pick):
    """6개 번호를 구간별 카운트 튜플로 변환."""
    counts = [0] * len(ZONES)
    for n in pick:
        counts[zone_index(n)] += 1
    return tuple(counts)


def _filter_history_by_allowed(history):
    """허용된 번호로만 구성된 회차만 남김 (제외 번호가 든 회차는 통계에서 제외)."""
    allowed_set = set(ALLOWED_NUMBERS)
    return [h for h in history if all(n in allowed_set for n in h)]


def zone_distribution(history):
    """과거 회차들을 구간별 카운트로 집계. 제외 번호가 든 회차는 자동으로 빠짐.

    반환:
      total_per_zone : Counter {zone_idx: 총 등장 횟수}
      pattern_counter: Counter {pattern_tuple: 회차 수}
    """
    total_per_zone = Counter()
    pattern_counter = Counter()
    for h in _filter_history_by_allowed(history):
        pat = zone_pattern(h)
        pattern_counter[pat] += 1
        for zi, c in enumerate(pat):
            total_per_zone[zi] += c
    return total_per_zone, pattern_counter


def adjacent_gaps(pick):
    """sorted 6개의 인접 간격 5개 리스트."""
    s = sorted(pick)
    return [b - a for a, b in zip(s, s[1:])]


def gap_distribution(history):
    """과거 회차의 인접 간격 빈도. 제외 번호가 든 회차는 자동으로 빠짐."""
    c = Counter()
    for h in _filter_history_by_allowed(history):
        c.update(adjacent_gaps(h))
    return c


def load_existing_picks(picks_dir: Path):
    """picks/ 안의 모든 xlsx 에서 이미 만들어진 6조합을 모아 set 으로 반환.

    n1..n6 컬럼이 있는 시트(버전별 시트, Combined 시트)만 인식한다.
    Frequency / Zones / Summary 등은 자동으로 무시된다.
    """
    existing = set()
    if not picks_dir.exists():
        return existing
    for xlsx_file in sorted(picks_dir.glob("*.xlsx")):
        # Excel 임시/잠금 파일(~$xxx.xlsx) 같은 것은 건너뜀
        if xlsx_file.name.startswith("~$") or xlsx_file.name.startswith("."):
            continue
        try:
            wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        except Exception as e:
            print(f"[경고] {xlsx_file.name} 로드 실패: {e}")
            continue
        try:
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue
                if not header:
                    continue
                try:
                    idx = [header.index(f"n{i}") for i in range(1, 7)]
                except ValueError:
                    continue
                for row in rows:
                    if not row:
                        continue
                    try:
                        nums = [int(row[i]) for i in idx if row[i] is not None]
                    except (TypeError, ValueError):
                        continue
                    if len(nums) != 6:
                        continue
                    pick = tuple(sorted(nums))
                    if len(set(pick)) == 6:
                        existing.add(pick)
        finally:
            wb.close()
    return existing


def load_history(path: Path):
    """xlsx에서 (번호1..번호6) 튜플 리스트를 로드. 보너스/회차/상금은 제외."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)
    history = []
    for row in rows:
        nums = row[1:7]
        if any(n is None for n in nums):
            continue
        history.append(tuple(int(n) for n in nums))
    wb.close()
    return history


def number_frequency(history):
    """각 번호(1~45)가 과거 1등 당첨에 등장한 횟수 Counter."""
    counter = Counter()
    for nums in history:
        counter.update(nums)
    return counter


def number_probability(counter):
    """각 허용 번호의 출현 확률 dict {n: p}. sum(p) == 1.0 (허용 번호 한정).

    제외된 번호는 분자/분모 모두에서 빠진다.
    """
    total = sum(counter.get(n, 0) for n in ALLOWED_NUMBERS)
    if total == 0:
        return {n: 0.0 for n in ALLOWED_NUMBERS}
    return {n: counter.get(n, 0) / total for n in ALLOWED_NUMBERS}


def build_weights(counter, mode):
    """샘플링용 가중치 리스트 (ALLOWED_NUMBERS 순서).

    mode:
      - 'weighted' : 출현 확률 그대로 (자주 나온 번호일수록 채택 가능성↑)
      - 'inverse'  : 최대확률-확률 (드물게 나온 번호 우대)
      - 'uniform'  : 모든 번호 동일 가중
    """
    probs = number_probability(counter)
    if mode == "uniform":
        return [1.0] * len(ALLOWED_NUMBERS)
    if mode == "weighted":
        return [probs[n] + 1e-6 for n in ALLOWED_NUMBERS]
    if mode == "inverse":
        max_p = max(probs.values()) if probs else 0
        return [(max_p + 1e-6) - probs[n] for n in ALLOWED_NUMBERS]
    raise ValueError(f"unknown mode: {mode}")


def count_consecutive_pairs(nums):
    s = sorted(nums)
    return sum(1 for a, b in zip(s, s[1:]) if b - a == 1)


def has_consecutive_run(pick, run_length=3):
    """3개 이상 연속된 숫자가 있으면 True (예: 1-2-3, 4-5-6-7)."""
    s = sorted(pick)
    run = 1
    for a, b in zip(s, s[1:]):
        if b - a == 1:
            run += 1
            if run >= run_length:
                return True
        else:
            run = 1
    return False


def weighted_sample_unique(weights, k):
    """ALLOWED_NUMBERS 에서 가중치 기반으로 k개 비복원 추출."""
    population = list(ALLOWED_NUMBERS)
    w = list(weights)
    picked = []
    for _ in range(k):
        total = sum(w)
        if total <= 0:
            break
        r = random.random() * total
        upto = 0.0
        for idx, weight in enumerate(w):
            upto += weight
            if upto >= r:
                picked.append(population[idx])
                w[idx] = 0
                break
    return tuple(sorted(picked))


def _weighted_pick_from_pool(pool, pool_weights, k):
    """주어진 후보 pool 에서 가중치 기반으로 k개 비복원 추출."""
    avail = list(pool)
    w = list(pool_weights)
    picked = []
    for _ in range(k):
        total = sum(w)
        if total <= 0:
            break
        r = random.random() * total
        upto = 0.0
        for idx, weight in enumerate(w):
            upto += weight
            if upto >= r:
                picked.append(avail[idx])
                w[idx] = 0
                break
    return picked


def zone_pattern_pick(counter, pattern_counter):
    """과거 zone 패턴 분포에서 한 패턴을 확률에 비례해 추첨한 뒤,
    각 zone 내부는 번호별 확률에 따라 비복원 추출하여 6개 번호 반환."""
    probs = number_probability(counter)
    patterns = list(pattern_counter.keys())
    weights = list(pattern_counter.values())
    pattern = random.choices(patterns, weights=weights, k=1)[0]
    pick = []
    for zi, count in enumerate(pattern):
        if count == 0:
            continue
        zone_nums = ZONES[zi][2]  # members
        zone_w = [probs[n] + 1e-9 for n in zone_nums]
        pick.extend(_weighted_pick_from_pool(zone_nums, zone_w, count))
    return tuple(sorted(pick))


def generate_pick(counter, history_set, mode, target_consecutive,
                  *, weights=None, pattern_counter=None, exclude_pairs=()):
    """모드별 단일 추첨. 과거 당첨/연속쌍/동시출현 금지 조건 만족할 때까지 재시도.

    exclude_pairs: list[frozenset[int]] — 각 set이 pick 안에 모두 포함되면 폐기.
    """
    pick_set_check = [frozenset(s) for s in exclude_pairs]
    for _ in range(MAX_ATTEMPTS):
        if mode == "zone":
            pick = zone_pattern_pick(counter, pattern_counter)
            if len(set(pick)) != PICK_SIZE:
                continue
        else:
            pick = weighted_sample_unique(weights, PICK_SIZE)
        if pick in history_set:
            continue
        if count_consecutive_pairs(pick) != target_consecutive:
            continue
        if has_consecutive_run(pick, 3):
            continue
        if pick_set_check:
            ps = set(pick)
            if any(p.issubset(ps) for p in pick_set_check):
                continue
        return pick
    raise RuntimeError(
        f"could not produce a pick (mode={mode}, target_consecutive={target_consecutive}) "
        f"after {MAX_ATTEMPTS} attempts"
    )


def generate_set(count, counter, history_set, mode, target_consecutive,
                 *, weights=None, pattern_counter=None, exclude_pairs=(),
                 seen=None):
    """count 개의 픽을 생성. seen 을 외부에서 넘기면 호출 간(=버전 간) 중복도 방지."""
    if seen is None:
        seen = set()
    picks = []
    while len(picks) < count:
        pick = generate_pick(counter, history_set, mode, target_consecutive,
                             weights=weights, pattern_counter=pattern_counter,
                             exclude_pairs=exclude_pairs)
        if pick in seen:
            continue
        seen.add(pick)
        picks.append(pick)
    return picks


VERSIONS = [
    ("hot-no-consecutive",     "weighted", 0, "고빈도 가중치 · 연속번호 없음"),
    ("hot-one-consecutive",    "weighted", 1, "고빈도 가중치 · 연속번호 1쌍"),
    ("hot-two-consecutive",    "weighted", 2, "고빈도 가중치 · 연속번호 2쌍"),
    ("cold-no-consecutive",    "inverse",  0, "저빈도 가중치 · 연속번호 없음"),
    ("cold-one-consecutive",   "inverse",  1, "저빈도 가중치 · 연속번호 1쌍"),
    ("uniform-no-consecutive", "uniform",  0, "균등 랜덤 · 연속번호 없음"),
    ("uniform-one-consecutive", "uniform", 1, "균등 랜덤 · 연속번호 1쌍"),
    ("zone-no-consecutive",    "zone",     0, "구간패턴 + 번호확률 · 연속번호 없음"),
    ("zone-one-consecutive",   "zone",     1, "구간패턴 + 번호확률 · 연속번호 1쌍"),
    ("zone-two-consecutive",   "zone",     2, "구간패턴 + 번호확률 · 연속번호 2쌍"),
]


def format_pick(pick):
    return " ".join(f"{n:2d}" for n in pick)


def print_zone_table(history):
    print("=" * 60)
    print("구간별 출현 분포 (1회차 = 6개 번호, 제외 번호 든 회차는 빠짐)")
    print("=" * 60)
    filtered = _filter_history_by_allowed(history)
    total_per_zone, pattern_counter = zone_distribution(history)
    n_rounds = len(filtered)
    total_picks = n_rounds * PICK_SIZE
    print(f"  유효 회차: {n_rounds}건  /  허용 번호: {len(ALLOWED_NUMBERS)}개  /  구간 {len(ZONES)}개")
    print("  구간       총등장   회차당평균    구간확률")
    for zi, (lo, hi, _members) in enumerate(ZONES):
        cnt = total_per_zone.get(zi, 0)
        per_round = cnt / n_rounds if n_rounds else 0
        prob = cnt / total_picks * 100 if total_picks else 0
        print(f"  {lo:2d}-{hi:2d}    {cnt:6d}      {per_round:5.3f}      {prob:5.2f}%")
    print(f"\n  자주 나오는 구간 패턴 Top 10  (전체 {len(pattern_counter)}종)")
    for pat, c in pattern_counter.most_common(10):
        print(f"    {pat}  : {c:4d}회 ({c / n_rounds * 100:5.2f}%)")
    print()


def print_gap_table(history):
    print("=" * 60)
    print("인접 번호 간격 분포 (sorted 6개의 차이 5개씩)")
    print("=" * 60)
    gaps = gap_distribution(history)
    total = sum(gaps.values())
    print("  간격   빈도    확률")
    items = sorted(gaps.items())
    for g, c in items:
        print(f"  {g:3d}   {c:5d}   {c / total * 100:5.2f}%")
    avg = sum(g * c for g, c in items) / total if total else 0
    print(f"\n  평균 인접 간격: {avg:.3f}")
    print()


def print_frequency_table(counter):
    print("=" * 60)
    print(f"번호별 출현 빈도 / 확률  (허용 번호 {len(ALLOWED_NUMBERS)}개 한정)")
    print("=" * 60)
    probs = number_probability(counter)
    rows = sorted(ALLOWED_NUMBERS, key=lambda n: (-counter.get(n, 0), n))
    for i, n in enumerate(rows, 1):
        f = counter.get(n, 0)
        p = probs[n] * 100
        end = "\n" if i % 5 == 0 else "   "
        print(f"{n:2d}: {f:3d}회 ({p:4.2f}%)", end=end)
    if len(rows) % 5:
        print()
    print(f"  [검증] 확률 합계 = {sum(probs.values()):.6f}  (1.000000 이어야 함)")
    if EXCLUDED_NUMBERS:
        print(f"  [제외] 후보에서 빠진 번호: {sorted(EXCLUDED_NUMBERS)}")
    print()


def verify_pipeline(history, history_set, counter, trials=20_000):
    """가중치 샘플링과 과거 당첨 제외 로직을 실측으로 검증."""
    print("=" * 60)
    print(f"[VERIFY] 샘플링 시뮬레이션 ({trials:,}회)")
    print("=" * 60)

    probs = number_probability(counter)

    # (1) weighted 모드: 경험확률과 시뮬레이션 빈도가 비슷해야 함
    weights = build_weights(counter, "weighted")
    sim = Counter()
    for _ in range(trials):
        sim.update(weighted_sample_unique(weights, PICK_SIZE))
    sim_total = sum(sim.values())
    print("  weighted 모드 상위 5개 번호 — 기대확률 vs 시뮬레이션확률")
    top5 = sorted(NUMBER_RANGE, key=lambda n: -probs[n])[:5]
    for n in top5:
        expected = probs[n] * 100
        observed = sim[n] / sim_total * 100
        print(f"    번호 {n:2d}: 기대 {expected:5.2f}%   관측 {observed:5.2f}%")

    # (2) 과거 당첨 제외 검증: 강제로 generate_pick 1만번 돌려도 history에 없어야
    print("\n  과거 당첨 제외 검증 (3,000세트 추첨)")
    fail = 0
    for _ in range(3_000):
        pick = generate_pick(counter, history_set, "weighted",
                             target_consecutive=0, weights=weights)
        if pick in history_set:
            fail += 1
    print(f"    과거 당첨 조합 중복 발생: {fail}건  (0건이어야 함)")

    # (3) sanity: 실제 첫 회차 당첨번호가 history_set에 들어있는가
    sample = tuple(sorted(history[0]))
    in_set = sample in history_set
    print(f"\n  과거 첫 행 {sample} 이 history_set 에 포함됨: {in_set}  (True 여야 함)")
    print()


def export_to_xlsx(path: Path, counter, results, history=None, include_combined=False):
    """추첨 결과를 엑셀로 저장.

    results: list of (key, label, mode, target_consec, picks)
    시트 구성:
      - Frequency    : 번호별 빈도/확률
      - Zones        : 구간별 출현/확률
      - ZonePatterns : 구간 카운트 패턴 빈도
      - Gaps         : 인접 번호 간격 분포
      - Summary      : 버전별 메타
      - Combined     : 종합 모드일 때, 모든 버전 결과 통합
      - <key>        : 버전별 추첨 세트
    """
    wb = openpyxl.Workbook()

    # 1) Frequency 시트 (허용 번호만)
    ws_freq = wb.active
    ws_freq.title = "Frequency"
    ws_freq.append(["번호", "출현횟수", "확률(%)"])
    probs = number_probability(counter)
    for n in ALLOWED_NUMBERS:
        ws_freq.append([n, counter.get(n, 0), round(probs[n] * 100, 4)])
    for col, width in zip("ABC", (8, 12, 12)):
        ws_freq.column_dimensions[col].width = width

    # 1-2) Zones / ZonePatterns / Gaps 시트
    if history:
        filtered = _filter_history_by_allowed(history)
        total_per_zone, pattern_counter = zone_distribution(history)
        n_rounds = len(filtered)
        total_picks = n_rounds * PICK_SIZE

        ws_zone = wb.create_sheet("Zones")
        ws_zone.append(["구간", "범위", "멤버수", "총등장", "회차당평균", "구간확률(%)"])
        for zi, (lo, hi, members) in enumerate(ZONES):
            cnt = total_per_zone.get(zi, 0)
            ws_zone.append([
                f"Z{zi+1}",
                f"{lo}-{hi}",
                len(members),
                cnt,
                round(cnt / n_rounds, 4) if n_rounds else 0,
                round(cnt / total_picks * 100, 4) if total_picks else 0,
            ])
        for col, width in zip("ABCDEF", (8, 10, 10, 10, 14, 14)):
            ws_zone.column_dimensions[col].width = width

        ws_pat = wb.create_sheet("ZonePatterns")
        zone_header = "패턴(" + ",".join(f"Z{i+1}" for i in range(len(ZONES))) + ")"
        ws_pat.append([zone_header, "회차수", "확률(%)"])
        for pat, c in pattern_counter.most_common():
            ws_pat.append([
                "(" + ", ".join(str(x) for x in pat) + ")",
                c,
                round(c / n_rounds * 100, 4),
            ])
        for col, width in zip("ABC", (24, 10, 12)):
            ws_pat.column_dimensions[col].width = width
        ws_pat.freeze_panes = "A2"

        ws_gap = wb.create_sheet("Gaps")
        ws_gap.append(["간격", "빈도", "확률(%)"])
        gaps = gap_distribution(history)
        gtotal = sum(gaps.values())
        for g in sorted(gaps):
            ws_gap.append([g, gaps[g], round(gaps[g] / gtotal * 100, 4)])
        for col, width in zip("ABC", (8, 10, 12)):
            ws_gap.column_dimensions[col].width = width

    # 2) Summary 시트
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["버전키", "설명", "가중치모드", "목표 연속쌍", "생성 세트수"])
    for key, label, mode, target_consec, picks in results:
        ws_sum.append([key, label, mode, target_consec, len(picks)])
    ws_sum.append([])
    ws_sum.append(["합계", "", "", "", sum(len(r[4]) for r in results)])
    for col, width in zip("ABCDE", (24, 36, 12, 12, 12)):
        ws_sum.column_dimensions[col].width = width

    # 3) Combined 시트 (종합 모드) — 행 자체도 n1..n6 키로 오름차순
    if include_combined:
        ws_all = wb.create_sheet("Combined")
        ws_all.append(["순번", "버전키", "n1", "n2", "n3", "n4", "n5", "n6", "합계", "연속쌍"])
        all_entries = []
        for key, label, mode, target_consec, picks in results:
            for pick in picks:
                sp = tuple(sorted(pick))
                all_entries.append((sp, key))
        all_entries.sort(key=lambda e: e[0])  # (n1, n2, ..., n6) 사전식 오름차순
        for i, (sp, key) in enumerate(all_entries, 1):
            ws_all.append([i, key, *sp, sum(sp), count_consecutive_pairs(sp)])
        for col, width in zip("ABCDEFGHIJ", (8, 24, 6, 6, 6, 6, 6, 6, 8, 10)):
            ws_all.column_dimensions[col].width = width
        ws_all.freeze_panes = "A2"

    # 4) 버전별 시트 — 행 자체도 n1..n6 키로 오름차순
    for key, label, mode, target_consec, picks in results:
        ws = wb.create_sheet(key[:31])  # 시트명 31자 제한
        ws.append(["세트번호", "n1", "n2", "n3", "n4", "n5", "n6", "합계", "연속쌍"])
        sorted_picks = sorted((tuple(sorted(p)) for p in picks))
        for i, sp in enumerate(sorted_picks, 1):
            ws.append([i, *sp, sum(sp), count_consecutive_pairs(sp)])
        for col, width in zip("ABCDEFGHI", (10, 6, 6, 6, 6, 6, 6, 8, 10)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="과거 당첨 데이터 기반 로또 번호 추첨")
    parser.add_argument("-c", "--count", type=int, default=5,
                        help="버전별로 생성할 추첨 세트 수 (기본: 5)")
    parser.add_argument("--versions", nargs="*", default=None,
                        help=f"실행할 버전 키 (생략시 전체). 가능한 키: {[v[0] for v in VERSIONS]}")
    parser.add_argument("--seed", type=int, default=None, help="랜덤 시드")
    parser.add_argument("--no-stats", action="store_true", help="번호별 빈도 출력 생략")
    parser.add_argument("--verify", action="store_true",
                        help="확률 가중치/과거 당첨 제외 로직을 시뮬레이션으로 검증")
    parser.add_argument("--xlsx", type=str, default="lotto", metavar="STEM",
                        help="저장 파일명 stem (기본 'lotto'). "
                             "실제 경로는 picks/{stem}__excl-{nums}__{YYYYMMDD-HHMMSS}.xlsx")
    parser.add_argument("--no-save", action="store_true",
                        help="엑셀 저장 비활성화 (기본: 항상 저장)")
    parser.add_argument("--warmup", type=int, default=0, metavar="N",
                        help="실제 출력 전 N개의 픽을 미리 만들어 중복 방지 풀(seen)에만 "
                             "추가하고 버린다. 결과적으로 실제 출력은 '워밍업 N개 + picks/ "
                             "기존 픽들' 의 여집합 안에서만 선택된다.")
    parser.add_argument("--repeat", type=int, default=1, metavar="N",
                        help="'추첨 + 저장' 한 묶음을 N번 반복한다. 매 라운드는 직전까지 "
                             "쌓인 모든 픽(이전 라운드 + picks/ 기존)을 피해 새 6조합만 뽑는다. "
                             "각 라운드는 별도의 xlsx 파일로 저장되며 파일명에 _r001, _r002, ... "
                             "라운드 인덱스가 붙는다. 마지막 파일이 곧 '나온 케이스를 모두 "
                             "제외한 여집합' 에서 뽑은 결과에 해당.")
    parser.add_argument("--final-count", type=int, default=None, metavar="N",
                        help="--repeat 와 함께 사용. 마지막 라운드(여집합 라운드)에서만 "
                             "총 N개를 뽑게 한다. 앞 라운드들은 --total/-c 의 카운트를 그대로 "
                             "쓰고, 마지막만 N개로 늘리거나 줄일 수 있다. "
                             "예: --total 50 --repeat 200 --final-count 1000")
    parser.add_argument("-t", "--total", type=int, default=None,
                        help="종합 모드: 선택된 모든 버전 합산 N개 (각 버전에 균등 분배). -c 무시")
    parser.add_argument("--zones", type=int, default=7,
                        help="구간 분할 개수 (기본 7, 권장 7~8). 1~45 가능")
    parser.add_argument("--exclude-pair", action="append", default=[], metavar="N1,N2[,N3...]",
                        help="해당 번호들이 모두 동시에 등장하면 폐기 "
                             "(예: --exclude-pair 1,2). 반복 지정 가능")
    parser.add_argument("--exclude-numbers", type=str, default=None, metavar="N1,N2,...",
                        help=f"후보에서 영구히 제외할 번호 (예: --exclude-numbers 1,2,45). "
                             f"미지정 시 기본 제외 {sorted(DEFAULT_EXCLUDED_NUMBERS)} 가 적용된다. "
                             f"빈도/확률/구간 분할도 제외 후 진행")
    args = parser.parse_args()

    # 후보 풀 갱신
    global EXCLUDED_NUMBERS, ALLOWED_NUMBERS, ZONES
    if args.exclude_numbers:
        try:
            EXCLUDED_NUMBERS = {int(x) for x in args.exclude_numbers.split(",") if x.strip()}
        except ValueError:
            print(f"[오류] --exclude-numbers 파싱 실패: {args.exclude_numbers!r}")
            return
        ALLOWED_NUMBERS = sorted(set(NUMBER_RANGE) - EXCLUDED_NUMBERS)
        print(f"[제외] 후보 번호 {len(EXCLUDED_NUMBERS)}개 제거: {sorted(EXCLUDED_NUMBERS)} "
              f"(남은 후보 {len(ALLOWED_NUMBERS)}개)")
    else:
        EXCLUDED_NUMBERS = set(DEFAULT_EXCLUDED_NUMBERS)
        ALLOWED_NUMBERS = sorted(set(NUMBER_RANGE) - EXCLUDED_NUMBERS)
        print(f"[기본 제외] 후보 번호 {len(EXCLUDED_NUMBERS)}개 제거: "
              f"{sorted(EXCLUDED_NUMBERS)} (남은 후보 {len(ALLOWED_NUMBERS)}개)")

    # 구간 분할 갱신
    ZONES = make_zones(args.zones, ALLOWED_NUMBERS)
    _refresh_zone_index()
    print(f"[구간] {len(ZONES)}구간: " +
          ", ".join(f"{lo}-{hi}({len(m)})" for lo, hi, m in ZONES))

    # exclude-pair 파싱
    exclude_pairs = []
    for s in args.exclude_pair:
        try:
            nums = frozenset(int(x) for x in s.split(","))
        except ValueError:
            print(f"[경고] --exclude-pair 파싱 실패: {s!r}")
            continue
        if len(nums) < 2:
            print(f"[경고] --exclude-pair 는 2개 이상이어야 함: {s!r}")
            continue
        exclude_pairs.append(nums)
    if exclude_pairs:
        print(f"[필터] 동시 등장 시 폐기할 조합: {[sorted(p) for p in exclude_pairs]}")

    if args.seed is not None:
        random.seed(args.seed)

    history = load_history(XLSX_PATH)
    history_set = set(tuple(sorted(h)) for h in history)
    counter = number_frequency(history)

    print(f"\n[로드 완료] 과거 회차 {len(history)}건  ·  history_set 크기 {len(history_set)}건")
    if not args.no_stats:
        print_frequency_table(counter)
        print_zone_table(history)
        print_gap_table(history)

    if args.verify:
        verify_pipeline(history, history_set, counter)
        return

    selected = VERSIONS
    if args.versions:
        keys = set(args.versions)
        selected = [v for v in VERSIONS if v[0] in keys]
        unknown = keys - {v[0] for v in VERSIONS}
        if unknown:
            print(f"[경고] 알 수 없는 버전 키 무시: {unknown}")

    # 버전별 카운트 결정: --total 지정시 균등 분배, 아니면 전부 args.count
    n_versions = len(selected)
    if args.total is not None:
        if n_versions == 0:
            print("[오류] 선택된 버전이 없습니다.")
            return
        base, rem = divmod(args.total, n_versions)
        per_version = [base + 1 if i < rem else base for i in range(n_versions)]
        print(f"[종합 모드] 총 {args.total}개를 {n_versions}개 버전에 분배: {per_version}")
    else:
        per_version = [args.count] * n_versions

    # zone 모드용 패턴 분포는 한 번만 계산
    _, pattern_counter = zone_distribution(history)

    # 모든 버전이 공유하는 픽 집합. 한 번 뽑힌 6조합은 다른 버전에서도 다시 등장하지 않음.
    # 또한 picks/ 안의 기존 xlsx 들에 이미 들어 있는 6조합도 미리 채워둬,
    # 파일이 여러 개여도 picks/ 전체에서 중복이 발생하지 않도록 한다.
    existing_picks = load_existing_picks(PICKS_DIR)
    if existing_picks:
        print(f"[기존] picks/ 안에 이미 {len(existing_picks)}건의 6조합 존재 — 중복 방지용으로 로드")
    global_seen = set(existing_picks)

    # 워밍업: --warmup N 이 주어지면 N개를 먼저 뽑아 seen 에만 넣고 버린다.
    # 출력은 (알고리즘 유니버스) − (워밍업 N개 ∪ picks/ 기존 픽) 의 여집합에서 선택됨.
    if args.warmup > 0:
        if n_versions == 0:
            print("[경고] --warmup 적용할 버전이 없음 — 워밍업 생략")
        else:
            wbase, wrem = divmod(args.warmup, n_versions)
            warmup_per_version = [wbase + 1 if i < wrem else wbase
                                  for i in range(n_versions)]
            before = len(global_seen)
            print("=" * 60)
            print(f"[워밍업] {args.warmup}개 사전 추첨 (저장/출력 안 함, seen 풀에만 추가)")
            print(f"         버전별 분배: {warmup_per_version}")
            print("=" * 60)
            for (key, mode, target_consec, label), n in zip(selected, warmup_per_version):
                if n <= 0:
                    continue
                weights = None if mode == "zone" else build_weights(counter, mode)
                try:
                    generate_set(n, counter, history_set, mode, target_consec,
                                 weights=weights, pattern_counter=pattern_counter,
                                 exclude_pairs=exclude_pairs, seen=global_seen)
                except RuntimeError as e:
                    print(f"  [warmup {key}] 생성 실패: {e}")
            added = len(global_seen) - before
            print(f"[워밍업 완료] seen 풀에 +{added}건 (총 {len(global_seen)}건 → 이후 출력은 이 여집합에서)\n")

    total_rounds = max(1, args.repeat)
    saved_paths = []
    seen_before_rounds = len(global_seen)
    quiet_picks = total_rounds > 1   # 라운드 모드일 땐 콘솔 미리보기 생략

    for round_idx in range(total_rounds):
        is_final_round = (round_idx == total_rounds - 1)
        # 마지막 라운드에서만 final_count 적용 (총량 → 버전별 균등 분배)
        if is_final_round and args.final_count is not None and n_versions > 0:
            fbase, frem = divmod(args.final_count, n_versions)
            this_per_version = [fbase + 1 if i < frem else fbase
                                for i in range(n_versions)]
        else:
            this_per_version = per_version

        if total_rounds > 1:
            print("\n" + "#" * 60)
            tag = "  [FINAL/여집합]" if is_final_round and args.final_count is not None else ""
            print(f"# Round {round_idx + 1} / {total_rounds}  "
                  f"(누적 seen: {len(global_seen):,}){tag}")
            if is_final_round and args.final_count is not None:
                print(f"#   목표 총 {args.final_count}개 (버전별 분배: {this_per_version})")
            print("#" * 60)

        results = []
        for (key, mode, target_consec, label), n in zip(selected, this_per_version):
            if n <= 0:
                continue
            weights = None if mode == "zone" else build_weights(counter, mode)
            if not quiet_picks:
                print("=" * 60)
                print(f"[{key}] {label}  ({n}세트)")
                print("=" * 60)
            try:
                picks = generate_set(n, counter, history_set, mode, target_consec,
                                     weights=weights, pattern_counter=pattern_counter,
                                     exclude_pairs=exclude_pairs, seen=global_seen)
            except RuntimeError as e:
                print(f"  [{key}] 생성 실패: {e}")
                continue
            if not quiet_picks:
                preview = picks if args.total is None else picks[:5]
                for i, pick in enumerate(preview, 1):
                    print(f"  {i:2d}. {format_pick(pick)}   (연속쌍: {count_consecutive_pairs(pick)})")
                if args.total is not None and len(picks) > len(preview):
                    print(f"  ... 외 {len(picks) - len(preview)}세트 (전체는 엑셀 또는 --total 미지정시 확인)")
                print()
            results.append((key, label, mode, target_consec, picks))

        round_total = sum(len(r[4]) for r in results)
        if args.total is not None and not quiet_picks:
            print(f"[종합] 총 {round_total}세트 생성 완료")

        if args.xlsx and not args.no_save:
            base_path = Path(args.xlsx).expanduser()
            stem = base_path.stem
            suffix = base_path.suffix or ".xlsx"
            excl_tag = ("-".join(str(n) for n in sorted(EXCLUDED_NUMBERS))
                        if EXCLUDED_NUMBERS else "none")
            ts = datetime.now().strftime("%Y%m%d-%H%M%S")
            round_tag = f"_r{round_idx + 1:03d}" if total_rounds > 1 else ""
            PICKS_DIR.mkdir(parents=True, exist_ok=True)
            out_path = PICKS_DIR / f"{stem}__excl-{excl_tag}__{ts}{round_tag}{suffix}"
            export_to_xlsx(out_path, counter, results, history=history,
                           include_combined=(args.total is not None))
            saved_paths.append(out_path)
            if quiet_picks:
                print(f"  → r{round_idx + 1:03d}: {round_total}세트 / 누적 seen {len(global_seen):,} / 저장: {out_path.name}")
            else:
                print(f"[엑셀 저장 완료] {out_path}")

    if total_rounds > 1:
        added = len(global_seen) - seen_before_rounds
        print("\n" + "=" * 60)
        print(f"[전체 완료] {total_rounds}라운드 / 신규 픽 {added:,}건 / 누적 seen {len(global_seen):,}건")
        if saved_paths:
            print(f"            저장된 파일: {len(saved_paths)}개 (마지막: {saved_paths[-1].name})")
            print(f"            마지막 파일이 '앞선 모든 라운드의 여집합' 에서 뽑힌 결과")
        print("=" * 60)


if __name__ == "__main__":
    main()
