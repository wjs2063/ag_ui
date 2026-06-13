"""
picks/ 폴더에 들어 있는 모든 엑셀파일의 기존 조합과 past_lotto.xlsx 의
역대 당첨번호를 모두 제외하고, 동일한 옵션(zone-*, 1/2/45 제외, 6구간)으로
500개의 새 조합을 생성한다. 결과는 picks/ 폴더에 실행시각 포함된 이름으로 저장.
"""

from datetime import datetime
from pathlib import Path

import openpyxl

import lotto_generator as lg

BASE_DIR = Path(__file__).parent
PICKS_DIR = BASE_DIR / "picks"
PAST_XLSX = BASE_DIR / "past_lotto.xlsx"

EXCLUDED_NUMBERS = {1, 2, 45}
N_ZONES = 6
TOTAL = 500
SELECTED_KEYS = ["zone-no-consecutive", "zone-one-consecutive", "zone-two-consecutive"]


def load_picks_from_dir(picks_dir: Path):
    """picks_dir 안의 모든 .xlsx 의 Combined 시트에서 6개 번호 조합 로드.

    반환: (picks_set, [(파일명, 조합 수), ...])
    """
    picks_set = set()
    file_stats = []
    for path in sorted(picks_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [경고] {path.name} 열기 실패: {e}")
            continue
        if "Combined" not in wb.sheetnames:
            wb.close()
            print(f"  [스킵] {path.name}: Combined 시트 없음")
            continue
        ws = wb["Combined"]
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nums = row[2:8]
            if any(x is None for x in nums):
                continue
            picks_set.add(tuple(sorted(int(x) for x in nums)))
            n += 1
        wb.close()
        file_stats.append((path.name, n))
    return picks_set, file_stats


def main():
    PICKS_DIR.mkdir(exist_ok=True)

    lg.EXCLUDED_NUMBERS = set(EXCLUDED_NUMBERS)
    lg.ALLOWED_NUMBERS = sorted(set(lg.NUMBER_RANGE) - EXCLUDED_NUMBERS)
    lg.ZONES = lg.make_zones(N_ZONES, lg.ALLOWED_NUMBERS)
    lg._refresh_zone_index()
    print(f"[제외 번호] {sorted(EXCLUDED_NUMBERS)}  ·  남은 후보 {len(lg.ALLOWED_NUMBERS)}개")
    print(f"[구간 {len(lg.ZONES)}개] " +
          ", ".join(f"{lo}-{hi}({len(m)})" for lo, hi, m in lg.ZONES))

    history = lg.load_history(PAST_XLSX)
    history_set = set(tuple(sorted(h)) for h in history)
    counter = lg.number_frequency(history)

    print(f"\n[picks 폴더] {PICKS_DIR}")
    existing, file_stats = load_picks_from_dir(PICKS_DIR)
    if file_stats:
        print(f"  로드한 파일 {len(file_stats)}개:")
        for name, n in file_stats:
            print(f"    - {name}  ({n}건)")
    else:
        print("  (비어 있음 — 처음 실행)")
    before = len(history_set)
    history_set.update(existing)
    added = len(history_set) - before
    print(f"[중복 제외 대상] 역대 당첨 {len(history)}건 + 기존 조합 {len(existing)}건 "
          f"(역대와 겹친 {len(existing) - added}건 포함)")

    selected = [v for v in lg.VERSIONS if v[0] in SELECTED_KEYS]
    n_versions = len(selected)
    base, rem = divmod(TOTAL, n_versions)
    per_version = [base + 1 if i < rem else base for i in range(n_versions)]
    print(f"[종합 모드] 총 {TOTAL}개 → {n_versions}개 버전에 분배: {per_version}")

    _, pattern_counter = lg.zone_distribution(history)

    results = []
    for (key, mode, target_consec, label), n in zip(selected, per_version):
        weights = None if mode == "zone" else lg.build_weights(counter, mode)
        print(f"  - [{key}] {n}세트 생성 중...")
        picks = lg.generate_set(
            n, counter, history_set, mode, target_consec,
            weights=weights, pattern_counter=pattern_counter,
        )
        # 다음 버전과의 중복도 방지
        history_set.update(picks)
        results.append((key, label, mode, target_consec, picks))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = PICKS_DIR / f"zone_only_500_{timestamp}.xlsx"
    lg.export_to_xlsx(out_path, counter, results, history=history, include_combined=True)

    total_generated = sum(len(r[4]) for r in results)
    print(f"[완료] {total_generated}세트 생성  ·  저장: {out_path.relative_to(BASE_DIR)}")


if __name__ == "__main__":
    main()
